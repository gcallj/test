"""
Gera summary_latest_v2.xlsx com TODOS os dados atualizados do base model
(fit 22.13, WR 72% mean). Inclui:

  Sheet "Apply_V2":
    Todas as colunas originais do Apply + novas colunas:
      - wr_target           (% alvo hits, NEW)
      - wr_target_robust    (% alvo OR time-exit lucrativo, NEW)
      - pct_exit_take/stop/time
      - alpha_ann_pp (full history)
      - rank_v2    (rank recalculado com win_rate_target 25% weight + outros ajustes)
      - confidence_v2

  Sheet "Summary_V2":
    Por ticker com todos os metricos de WR e distribucao

  Sheet "Win_Rate_Analysis":
    Comparacao WR atual vs WR_target por ticker
"""
import os
import sys
import json
import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO))
os.chdir(REPO)

os.environ.setdefault("GA_RUN_MODE", "load")
os.environ["GA_STAGE_GATE"] = "off"
os.environ["GA_ALPHA_FOCUS"] = "off"

import ga_run  # noqa: E402
from ga_run import (  # noqa: E402
    decode_global_params,
    backtest_stats_global_intraday,
    buyhold_capped,
    GA_MEMMAP_DIR,
)
from payload_store import PayloadStore  # noqa: E402


candidates = [
    os.path.abspath(GA_MEMMAP_DIR),
    os.path.abspath("output/ga_memmap_staged"),
]
store_path = None
for p in candidates:
    if os.path.exists(os.path.join(p, "meta.json")):
        store_path = p
        break
print(f"[INIT] store: {store_path}", flush=True)
store = PayloadStore(store_path, mode="r")

print(f"[INIT] Building ticker cache...", flush=True)
cache = {}
for tk in store.tickers:
    a = store.get_ticker_arrays(tk)
    p = {}
    for k in ("open", "high", "low", "close", "atr", "ma", "vol_rank", "volume",
              "ret_fwd", "long_votes", "short_votes"):
        if k in a and a[k] is not None:
            p[k] = np.asarray(a[k], dtype=np.float64)
    if "score_matrix" in a and a["score_matrix"] is not None:
        p["score_matrix"] = np.asarray(a["score_matrix"], dtype=np.float64)
    if p.get("close") is not None and len(p["close"]) >= 252:
        cache[tk] = p
print(f"[INIT] Cache: {len(cache)} tickers", flush=True)

# Load base checkpoint
with open("global_ga_checkpoint.json") as f:
    ck = json.load(f)
base_genome = ck["genome"]
gp = decode_global_params(base_genome)
print(f"[BASE] checkpoint fit={ck['fitness']:.3f}", flush=True)

# Compute per-ticker metrics with new WR metrics
print(f"\n[EVAL] Running backtest for all tickers...", flush=True)
per_ticker = {}
for tk, p in cache.items():
    st = backtest_stats_global_intraday(
        p["open"], p["high"], p["low"], p["close"],
        p["score_matrix"], p["atr"], gp, precomputed=p,
    )
    bh = buyhold_capped(p["close"])
    n_years = len(p["close"]) / 252.0
    ret_ann = (1 + st["total_return"]) ** (1 / max(n_years, 0.1)) - 1
    bh_ann = (1 + bh) ** (1 / max(n_years, 0.1)) - 1
    alpha_ann = ret_ann - bh_ann
    per_ticker[tk] = {
        "n_trades": int(st.get("n_trades", 0)),
        "n_years": round(n_years, 1),
        "win_rate_full": round(st.get("win_rate", 0) * 100, 2),
        "win_rate_target": round(st.get("win_rate_target", 0) * 100, 2),
        "win_rate_target_robust": round(st.get("win_rate_target_robust", 0) * 100, 2),
        "pct_exit_take": round(st.get("pct_exit_take", 0) * 100, 2),
        "pct_exit_stop": round(st.get("pct_exit_stop", 0) * 100, 2),
        "pct_exit_time": round(st.get("pct_exit_time", 0) * 100, 2),
        "ret_total_15a": round(st.get("total_return", 0), 4),
        "ret_ann_pct": round(ret_ann * 100, 2),
        "bh_ann_pct": round(bh_ann * 100, 2),
        "alpha_ann_pp": round(alpha_ann * 100, 2),
        "mdd_full_pct": round(abs(st.get("mdd", 0)) * 100, 2),
        "sharpe_full": round(st.get("sharpe", 0), 3),
        "big_wins_pct": round(st.get("big_wins_pct", 0) * 100, 2),
        "big_losses_pct": round(st.get("big_losses_pct", 0) * 100, 2),
        "dist_quality": round(st.get("dist_quality", 0), 3),
        "pct_gt5": round((st.get("pct_gt15", 0) + st.get("pct_10_15", 0) + st.get("pct_5_10", 0)) * 100, 2),
        "avg_trade_pct": round(st.get("avg_trade", 0) * 100, 3),
    }

# Load original xlsx and merge
print(f"\n[BUILD] Loading original summary_latest.xlsx", flush=True)
wb = load_workbook("summary_latest.xlsx")

# Read Apply sheet
ws_app = wb["Apply"]
app_headers = [c.value for c in next(ws_app.rows)]
apply_rows = []
for row in ws_app.iter_rows(min_row=2, values_only=True):
    apply_rows.append(dict(zip(app_headers, row)))
apply_df = pd.DataFrame(apply_rows)

# Add new metrics per ticker
def lookup(tk, key, default=None):
    return per_ticker.get(tk, {}).get(key, default)

apply_df["wr_target"] = apply_df["ticker"].map(lambda t: lookup(t, "win_rate_target"))
apply_df["wr_target_robust"] = apply_df["ticker"].map(lambda t: lookup(t, "win_rate_target_robust"))
apply_df["pct_exit_take"] = apply_df["ticker"].map(lambda t: lookup(t, "pct_exit_take"))
apply_df["pct_exit_stop"] = apply_df["ticker"].map(lambda t: lookup(t, "pct_exit_stop"))
apply_df["pct_exit_time"] = apply_df["ticker"].map(lambda t: lookup(t, "pct_exit_time"))
apply_df["alpha_ann_pp_15a"] = apply_df["ticker"].map(lambda t: lookup(t, "alpha_ann_pp"))
apply_df["ret_ann_15a"] = apply_df["ticker"].map(lambda t: lookup(t, "ret_ann_pct"))
apply_df["bh_ann_15a"] = apply_df["ticker"].map(lambda t: lookup(t, "bh_ann_pct"))
apply_df["mdd_full_15a"] = apply_df["ticker"].map(lambda t: lookup(t, "mdd_full_pct"))
apply_df["n_trades_15a"] = apply_df["ticker"].map(lambda t: lookup(t, "n_trades"))

# --- Coherence test: recompute rank_v2 with WR weighted more strongly ---
# Current weights: confidence 20%, upside 20%, timing 15%, pivot 15%, rr 10%,
# win_rate 10%, regime 5%, stop 5% = 100%
# New weights:     confidence 15%, upside 15%, timing 12%, pivot 12%, rr 8%,
# win_rate 30% (DOUBLE) + wr_target 5% (NEW), regime 3%
def compute_rank_v2(row):
    conf = float(row.get("confidence", 50))
    wr = float(row.get("win_rate", 50))
    wrt = float(row.get("wr_target") or 0)
    rr = float(row.get("RR", 1.5)) if row.get("RR") is not None else 1.5
    # Approximate upside/timing/pivot from existing columns (not perfect)
    # Use potencial_pct as proxy for upside
    pot_raw = row.get("potencial_pct", "0%")
    try:
        pot = float(str(pot_raw).replace("+", "").replace("%", "").strip())
    except Exception:
        pot = 0.0
    upside_proxy = np.clip(pot * 3, 0, 100)  # +10% = 30 score, +33% = 100
    # Approximate rank_v2 = heavier WR, include WR_target
    rank_v2 = (
        0.15 * conf +
        0.15 * upside_proxy +
        0.30 * wr +             # <- BOOSTED from 10% to 30%
        0.05 * wrt +            # <- NEW: wr_target contribution
        0.10 * min(rr / 3.0, 1.0) * 100 +
        0.15 * 50 +   # placeholder for timing/pivot we can't easily reconstruct
        0.10 * 50     # placeholder for other factors
    )
    return round(rank_v2, 1)

def compute_conf_v2(row):
    """Confidence v2: boost q_wr weight from 18% to 28% and add q_wr_target 10%."""
    base_conf = float(row.get("confidence", 50))
    wr = float(row.get("win_rate", 50))
    wrt = float(row.get("wr_target") or 0)
    # Weighted blend: 70% of base + 20% direct WR + 10% WR_target
    return round(0.70 * base_conf + 0.20 * wr + 0.10 * wrt, 1)

apply_df["rank_v2"] = apply_df.apply(compute_rank_v2, axis=1)
apply_df["confidence_v2"] = apply_df.apply(compute_conf_v2, axis=1)

# Sort by rank_v2 descending
apply_df = apply_df.sort_values("rank_v2", ascending=False)

# Add Apply_V2 sheet
if "Apply_V2" in wb.sheetnames:
    del wb["Apply_V2"]
ws_app2 = wb.create_sheet("Apply_V2")
ws_app2.append(list(apply_df.columns))
for _, row in apply_df.iterrows():
    ws_app2.append([row[c] if pd.notna(row[c]) else None for c in apply_df.columns])

# Add Summary_V2 sheet
summary_rows = []
for tk, m in per_ticker.items():
    row = {"ticker": tk}
    row.update(m)
    summary_rows.append(row)
summary_df = pd.DataFrame(summary_rows).sort_values("win_rate_full", ascending=False)

if "Summary_V2" in wb.sheetnames:
    del wb["Summary_V2"]
ws_sum2 = wb.create_sheet("Summary_V2")
ws_sum2.append(list(summary_df.columns))
for _, row in summary_df.iterrows():
    ws_sum2.append([row[c] if pd.notna(row[c]) else None for c in summary_df.columns])

# Add Win_Rate_Analysis sheet: side-by-side comparison
wra_rows = []
for tk, m in per_ticker.items():
    wra_rows.append({
        "ticker": tk,
        "n_trades": m["n_trades"],
        "WR_full (tr>0)": m["win_rate_full"],
        "WR_target (alvo hit)": m["win_rate_target"],
        "delta_pp": round(m["win_rate_full"] - m["win_rate_target"], 2),
        "pct_exit_take": m["pct_exit_take"],
        "pct_exit_stop": m["pct_exit_stop"],
        "pct_exit_time": m["pct_exit_time"],
        "ret_ann_pct": m["ret_ann_pct"],
        "alpha_ann_pp": m["alpha_ann_pp"],
    })
wra_df = pd.DataFrame(wra_rows).sort_values("WR_full (tr>0)", ascending=False)

if "Win_Rate_Analysis" in wb.sheetnames:
    del wb["Win_Rate_Analysis"]
ws_wra = wb.create_sheet("Win_Rate_Analysis")
ws_wra.append(list(wra_df.columns))
for _, row in wra_df.iterrows():
    ws_wra.append([row[c] if pd.notna(row[c]) else None for c in wra_df.columns])

# Save
wb.save("summary_latest_v2.xlsx")
print(f"[SAVE] summary_latest_v2.xlsx", flush=True)

# Print coherence comparison: rank_v2 vs rank, with correlations
import numpy as np
rank_arr = apply_df["rank"].astype(float).values
rank_v2_arr = apply_df["rank_v2"].astype(float).values
wr_arr = apply_df["win_rate"].astype(float).values
wrt_arr = apply_df["wr_target"].astype(float).values
conf_arr = apply_df["confidence"].astype(float).values
conf_v2_arr = apply_df["confidence_v2"].astype(float).values


def corr(a, b):
    mask = np.isfinite(a) & np.isfinite(b)
    if mask.sum() < 3:
        return np.nan
    return float(np.corrcoef(a[mask], b[mask])[0, 1])


print(f"\n=== COHERENCE COMPARISON (rank_v2 boost WR weight) ===", flush=True)
print(f"  rank    vs win_rate:       {corr(rank_arr, wr_arr):+.3f}", flush=True)
print(f"  rank_v2 vs win_rate:       {corr(rank_v2_arr, wr_arr):+.3f}", flush=True)
print(f"  rank    vs wr_target:      {corr(rank_arr, wrt_arr):+.3f}", flush=True)
print(f"  rank_v2 vs wr_target:      {corr(rank_v2_arr, wrt_arr):+.3f}", flush=True)
print(f"  conf    vs win_rate:       {corr(conf_arr, wr_arr):+.3f}", flush=True)
print(f"  conf_v2 vs win_rate:       {corr(conf_v2_arr, wr_arr):+.3f}", flush=True)
print(f"  conf    vs wr_target:      {corr(conf_arr, wrt_arr):+.3f}", flush=True)
print(f"  conf_v2 vs wr_target:      {corr(conf_v2_arr, wrt_arr):+.3f}", flush=True)

# Buy subset only
buys = apply_df[apply_df["signal"] == "buy"]
if len(buys) > 3:
    b_rank = buys["rank"].astype(float).values
    b_rank_v2 = buys["rank_v2"].astype(float).values
    b_wr = buys["win_rate"].astype(float).values
    b_wrt = buys["wr_target"].astype(float).values
    print(f"\n  BUY subset (n={len(buys)}):", flush=True)
    print(f"    rank    vs win_rate:  {corr(b_rank, b_wr):+.3f}", flush=True)
    print(f"    rank_v2 vs win_rate:  {corr(b_rank_v2, b_wr):+.3f}", flush=True)
    print(f"    rank    vs wr_target: {corr(b_rank, b_wrt):+.3f}", flush=True)
    print(f"    rank_v2 vs wr_target: {corr(b_rank_v2, b_wrt):+.3f}", flush=True)

print(f"\n=== TOP 10 BUYS BY rank_v2 ===", flush=True)
top_buys = buys.head(10) if len(buys) > 0 else apply_df.head(10)
for _, r in top_buys.iterrows():
    print(f"  {r['ticker']:<12} signal={r['signal']:<6} "
          f"rank={r.get('rank')} rank_v2={r.get('rank_v2'):>5.1f} "
          f"conf={r.get('confidence')} conf_v2={r.get('confidence_v2'):>5.1f} "
          f"WR={r.get('win_rate')} WR_tgt={r.get('wr_target'):>5.1f} "
          f"RR={r.get('RR')}",
          flush=True)
