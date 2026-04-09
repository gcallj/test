"""Inspect rank and confidence columns coherence with other metrics."""
import os
import sys
import json
from pathlib import Path
from openpyxl import load_workbook

REPO = Path(__file__).resolve().parent.parent
os.chdir(REPO)

xlsx = "summary_latest.xlsx"
wb = load_workbook(xlsx, read_only=True)
print("Sheets:", wb.sheetnames)

ws = wb["Apply"]
headers = [c.value for c in next(ws.rows)]
print(f"\n--- Apply columns ({len(headers)}) ---")
for i, h in enumerate(headers):
    print(f"  [{i}] {h}")

# Extract all rows
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    rows.append(dict(zip(headers, row)))

print(f"\n--- {len(rows)} rows ---")

# Look at rank, confidence, signal, win_rate, potencial_pct
cols_of_interest = ["ticker", "signal", "estagio", "rank", "confidence",
                    "win_rate", "potencial_pct", "queda_max", "RR", "regime"]

# Print top 10 by rank
print("\n=== TOP 10 BY RANK ===")
def rank_val(r):
    v = r.get("rank")
    try:
        return float(v) if v is not None else -1
    except Exception:
        return -1

sorted_by_rank = sorted(rows, key=lambda r: -rank_val(r))
for r in sorted_by_rank[:10]:
    print(f"  {r.get('ticker'):<12} signal={r.get('signal'):<6} "
          f"rank={r.get('rank')} conf={r.get('confidence')} "
          f"wr={r.get('win_rate')} pot={r.get('potencial_pct')} "
          f"RR={r.get('RR')} regime={r.get('regime')}")

# Print bottom 10
print("\n=== BOTTOM 10 BY RANK ===")
for r in sorted_by_rank[-10:]:
    print(f"  {r.get('ticker'):<12} signal={r.get('signal'):<6} "
          f"rank={r.get('rank')} conf={r.get('confidence')} "
          f"wr={r.get('win_rate')} pot={r.get('potencial_pct')} "
          f"RR={r.get('RR')} regime={r.get('regime')}")

# Print top 10 by confidence
print("\n=== TOP 10 BY CONFIDENCE ===")
def conf_val(r):
    v = r.get("confidence")
    try:
        return float(v) if v is not None else -1
    except Exception:
        return -1

sorted_by_conf = sorted(rows, key=lambda r: -conf_val(r))
for r in sorted_by_conf[:10]:
    print(f"  {r.get('ticker'):<12} signal={r.get('signal'):<6} "
          f"rank={r.get('rank')} conf={r.get('confidence')} "
          f"wr={r.get('win_rate')} pot={r.get('potencial_pct')} "
          f"RR={r.get('RR')} regime={r.get('regime')}")

# Check correlations
import numpy as np
def to_num(v):
    try:
        return float(v) if v is not None else np.nan
    except Exception:
        return np.nan

def col_arr(key):
    return np.array([to_num(r.get(key)) for r in rows])

rank_arr = col_arr("rank")
conf_arr = col_arr("confidence")
wr_arr = col_arr("win_rate")

def pot_to_num(v):
    if v is None:
        return np.nan
    s = str(v).replace("+", "").replace("%", "").strip()
    try:
        return float(s)
    except Exception:
        return np.nan

pot_arr = np.array([pot_to_num(r.get("potencial_pct")) for r in rows])
rr_arr = col_arr("RR")
regime_arr = col_arr("regime")

def corr(a, b):
    mask = np.isfinite(a) & np.isfinite(b)
    if mask.sum() < 3:
        return np.nan
    return float(np.corrcoef(a[mask], b[mask])[0, 1])

print(f"\n=== CORRELATIONS (Pearson) ===")
print(f"rank vs confidence:   {corr(rank_arr, conf_arr):+.3f}")
print(f"rank vs win_rate:     {corr(rank_arr, wr_arr):+.3f}")
print(f"rank vs potencial:    {corr(rank_arr, pot_arr):+.3f}")
print(f"rank vs RR:           {corr(rank_arr, rr_arr):+.3f}")
print(f"rank vs regime:       {corr(rank_arr, regime_arr):+.3f}")
print(f"confidence vs win_rate: {corr(conf_arr, wr_arr):+.3f}")
print(f"confidence vs potencial:{corr(conf_arr, pot_arr):+.3f}")
print(f"confidence vs RR:     {corr(conf_arr, rr_arr):+.3f}")
print(f"confidence vs regime: {corr(conf_arr, regime_arr):+.3f}")
print(f"win_rate vs potencial:{corr(wr_arr, pot_arr):+.3f}")

print(f"\n=== STATS ===")
def stats(name, arr):
    mask = np.isfinite(arr)
    if mask.sum() == 0:
        print(f"  {name:<12}: all nan ({len(arr)} rows)")
        return
    a = arr[mask]
    print(f"  {name:<12}: n={mask.sum():>3}/{len(arr)} "
          f"min={a.min():>8.3f} med={np.median(a):>8.3f} "
          f"mean={a.mean():>8.3f} max={a.max():>8.3f}")

stats("rank", rank_arr)
stats("confidence", conf_arr)
stats("win_rate", wr_arr)
stats("potencial", pot_arr)
stats("RR", rr_arr)
stats("regime", regime_arr)

# Filter BUY only
buys = [r for r in rows if r.get("signal") == "buy"]
print(f"\n=== BUY signals only ({len(buys)}) ===")
b_rank = np.array([to_num(r.get("rank")) for r in buys])
b_conf = np.array([to_num(r.get("confidence")) for r in buys])
b_wr = np.array([to_num(r.get("win_rate")) for r in buys])
b_pot = np.array([pot_to_num(r.get("potencial_pct")) for r in buys])
b_rr = np.array([to_num(r.get("RR")) for r in buys])

print(f"rank vs confidence:     {corr(b_rank, b_conf):+.3f}")
print(f"rank vs win_rate:       {corr(b_rank, b_wr):+.3f}")
print(f"rank vs potencial_pct:  {corr(b_rank, b_pot):+.3f}")
print(f"rank vs RR:             {corr(b_rank, b_rr):+.3f}")
print(f"confidence vs win_rate: {corr(b_conf, b_wr):+.3f}")
print(f"confidence vs potencial:{corr(b_conf, b_pot):+.3f}")

wb.close()
