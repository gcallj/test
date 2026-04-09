"""
Deep dive sobre win_rate do GA.

Responde:
1. Win rate considera somente positivos ou so trades que atingiram alvo?
   R: Atualmente win_rate = (tr > 0).mean() — TRADES POSITIVOS, nao so alvo.
2. Quanto do win_rate vem de take-profit real (alvo) vs outros (time-stop com lucro)?
3. Qual a distribuicao de exit types por ticker?

Alem de reportar, gera planilha output/summary_latest_v2.xlsx com:
  - win_rate                 (% tr > 0, atual)
  - win_rate_target          (% hit take-profit, NOVO)
  - win_rate_target_robust   (% hit take OU time-stop lucrativo)
  - pct_exit_take, pct_exit_stop, pct_exit_time
"""
import os
import sys
import json
import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO))
os.chdir(REPO)

os.environ.setdefault("GA_RUN_MODE", "load")
os.environ["GA_STAGE_GATE"] = "off"
os.environ["GA_ALPHA_FOCUS"] = "off"

import ga_run  # noqa: E402
from ga_run import (  # noqa: E402
    decode_global_params,
    backtest_stats_global_intraday,
    buyhold_capped,
    GA_MEMMAP_DIR,
)
from payload_store import PayloadStore  # noqa: E402


candidates = [
    os.path.abspath(GA_MEMMAP_DIR),
    os.path.abspath("output/ga_memmap_staged"),
]
store_path = None
for p in candidates:
    if os.path.exists(os.path.join(p, "meta.json")):
        store_path = p
        break
print(f"[INIT] store: {store_path}", flush=True)
store = PayloadStore(store_path, mode="r")

print(f"[INIT] Building ticker cache...", flush=True)
cache = {}
for tk in store.tickers:
    a = store.get_ticker_arrays(tk)
    p = {}
    for k in ("open", "high", "low", "close", "atr", "ma", "vol_rank", "volume",
              "ret_fwd", "long_votes", "short_votes"):
        if k in a and a[k] is not None:
            p[k] = np.asarray(a[k], dtype=np.float64)
    if "score_matrix" in a and a["score_matrix"] is not None:
        p["score_matrix"] = np.asarray(a["score_matrix"], dtype=np.float64)
    if p.get("close") is not None and len(p["close"]) >= 252:
        cache[tk] = p
print(f"[INIT] Cache: {len(cache)} tickers", flush=True)


# Load best checkpoint (fit 22.13)
with open("global_ga_checkpoint.json") as f:
    ck = json.load(f)
base_genome = ck["genome"]
gp = decode_global_params(base_genome)
print(f"[BASE] checkpoint fit {ck['fitness']:.3f}", flush=True)
print(f"[BASE] GA params: rr_ratio={gp.reward_risk_ratio} "
      f"stop_atr_mult={gp.stop_atr_mult} "
      f"partial_take_pct={gp.partial_take_pct} "
      f"trailing_stop_mode={gp.trailing_stop_mode} "
      f"time_stop_bars={gp.time_stop_bars}", flush=True)

# Run backtest per-ticker with new metrics
print(f"\n[EVAL] Running backtest with exit-type tracking...", flush=True)
rows = []
for tk, p in cache.items():
    st = backtest_stats_global_intraday(
        p["open"], p["high"], p["low"], p["close"],
        p["score_matrix"], p["atr"], gp, precomputed=p,
    )
    bh = buyhold_capped(p["close"])
    n_years = len(p["close"]) / 252.0
    ret_ann = (1 + st["total_return"]) ** (1 / max(n_years, 0.1)) - 1
    bh_ann = (1 + bh) ** (1 / max(n_years, 0.1)) - 1
    alpha_ann = ret_ann - bh_ann
    rows.append({
        "ticker": tk,
        "n_trades": int(st.get("n_trades", 0)),
        "n_years": round(n_years, 1),
        "win_rate": round(st.get("win_rate", 0) * 100, 2),
        "win_rate_target": round(st.get("win_rate_target", 0) * 100, 2),
        "win_rate_target_robust": round(st.get("win_rate_target_robust", 0) * 100, 2),
        "pct_exit_take": round(st.get("pct_exit_take", 0) * 100, 2),
        "pct_exit_stop": round(st.get("pct_exit_stop", 0) * 100, 2),
        "pct_exit_time": round(st.get("pct_exit_time", 0) * 100, 2),
        "ret_total": round(st["total_return"], 4),
        "ret_ann_pct": round(ret_ann * 100, 2),
        "bh_ann_pct": round(bh_ann * 100, 2),
        "alpha_ann_pp": round(alpha_ann * 100, 2),
        "mdd_pct": round(abs(st.get("mdd", 0)) * 100, 2),
        "sharpe": round(st.get("sharpe", 0), 3),
        "big_wins_pct": round(st.get("big_wins_pct", 0) * 100, 2),
        "big_losses_pct": round(st.get("big_losses_pct", 0) * 100, 2),
    })

df = pd.DataFrame(rows).sort_values("win_rate", ascending=False)

# Aggregate stats
print(f"\n=== AGGREGATE WIN RATE COMPARISONS ===", flush=True)
print(f"n tickers: {len(df)}", flush=True)
print(f"Metric                     median     mean     min     max", flush=True)
for col in ("win_rate", "win_rate_target", "win_rate_target_robust",
            "pct_exit_take", "pct_exit_stop", "pct_exit_time"):
    med = df[col].median()
    mn = df[col].mean()
    mi = df[col].min()
    mx = df[col].max()
    print(f"  {col:<24} {med:>7.2f}% {mn:>7.2f}% {mi:>7.2f}% {mx:>7.2f}%",
          flush=True)

# How much of the "WR = 72%" is actually take-profit vs other?
print(f"\n=== WIN RATE BREAKDOWN ===", flush=True)
print(f"Current GA win_rate median:       {df['win_rate'].median():.2f}%", flush=True)
print(f"  de onde vem:", flush=True)
print(f"    take-profit (alvo hit):       {df['pct_exit_take'].median():.2f}% (estrito)", flush=True)
print(f"    take + time-exit lucrativo:   {df['win_rate_target_robust'].median():.2f}% (robusto)", flush=True)
print(f"  delta (trades positivos mas nao alvo) = WR - WR_target", flush=True)
df["wr_non_target_pp"] = df["win_rate"] - df["win_rate_target"]
df["wr_non_target_robust_pp"] = df["win_rate"] - df["win_rate_target_robust"]
print(f"  delta (estrito):  {df['wr_non_target_pp'].median():.2f}pp", flush=True)
print(f"  delta (robusto):  {df['wr_non_target_robust_pp'].median():.2f}pp", flush=True)

print(f"\n=== TOP 10 TICKERS POR WIN_RATE (atual) ===", flush=True)
cols_show = ["ticker", "n_trades", "win_rate", "win_rate_target",
             "win_rate_target_robust", "pct_exit_take", "pct_exit_stop",
             "pct_exit_time", "ret_ann_pct", "alpha_ann_pp"]
for _, r in df.head(10).iterrows():
    print(f"  {r['ticker']:<12} n={int(r['n_trades']):>3} "
          f"WR={r['win_rate']:>5.1f}% "
          f"WR_tgt={r['win_rate_target']:>5.1f}% "
          f"WR_tgtR={r['win_rate_target_robust']:>5.1f}% "
          f"take={r['pct_exit_take']:>5.1f}% "
          f"stop={r['pct_exit_stop']:>5.1f}% "
          f"time={r['pct_exit_time']:>5.1f}% "
          f"ret={r['ret_ann_pct']:>+6.2f}% "
          f"alpha={r['alpha_ann_pp']:>+6.2f}pp",
          flush=True)

# Save CSV
df.to_csv("analysis/win_rate_deep_dive.csv", index=False)
print(f"\n[SAVE] analysis/win_rate_deep_dive.csv", flush=True)

# Build new summary xlsx by COPYING summary_latest.xlsx and adding NEW win_rate columns
src_xlsx = "summary_latest.xlsx"
dst_xlsx = "summary_latest_v2.xlsx"

# Read existing sheets
wb = load_workbook(src_xlsx)

# Add new Summary_V2 sheet with all rows including new metrics
from openpyxl import Workbook
if "Summary_V2" in wb.sheetnames:
    del wb["Summary_V2"]
ws_new = wb.create_sheet("Summary_V2")

# Write header
header = list(df.columns)
ws_new.append(header)
for _, row in df.iterrows():
    ws_new.append([row[c] for c in header])

# Update the existing Apply sheet to add new WR columns (merge on ticker)
# Load Apply sheet into a dataframe
ws_app = wb["Apply"]
apply_rows = []
app_headers = [c.value for c in next(ws_app.rows)]
for row in ws_app.iter_rows(min_row=2, values_only=True):
    apply_rows.append(dict(zip(app_headers, row)))
apply_df = pd.DataFrame(apply_rows)

# Merge with new metrics
merge_cols = ["ticker", "win_rate_target", "win_rate_target_robust",
              "pct_exit_take", "pct_exit_stop", "pct_exit_time", "alpha_ann_pp"]
extra_df = df[merge_cols].rename(columns={"win_rate_target": "wr_target",
                                          "win_rate_target_robust": "wr_target_robust"})
merged = apply_df.merge(extra_df, on="ticker", how="left")

# Create new Apply_V2 sheet
if "Apply_V2" in wb.sheetnames:
    del wb["Apply_V2"]
ws_app2 = wb.create_sheet("Apply_V2")
ws_app2.append(list(merged.columns))
for _, row in merged.iterrows():
    ws_app2.append([row[c] for c in merged.columns])

wb.save(dst_xlsx)
print(f"[SAVE] {dst_xlsx}", flush=True)

# Save JSON for downstream
summary = {
    "checkpoint_fit": ck["fitness"],
    "n_tickers": len(df),
    "win_rate_median": float(df["win_rate"].median()),
    "win_rate_target_median": float(df["win_rate_target"].median()),
    "win_rate_target_robust_median": float(df["win_rate_target_robust"].median()),
    "pct_exit_take_median": float(df["pct_exit_take"].median()),
    "pct_exit_stop_median": float(df["pct_exit_stop"].median()),
    "pct_exit_time_median": float(df["pct_exit_time"].median()),
    "delta_wr_strict_pp_median": float(df["wr_non_target_pp"].median()),
    "delta_wr_robust_pp_median": float(df["wr_non_target_robust_pp"].median()),
}
with open("analysis/win_rate_deep_dive.json", "w") as f:
    json.dump(summary, f, indent=2)
print(f"[SAVE] analysis/win_rate_deep_dive.json", flush=True)
