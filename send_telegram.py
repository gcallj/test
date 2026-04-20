"""Envia summary_latest.xlsx atual para o Telegram com caption completo."""
import os
import json
import numpy as np
import pandas as pd
from openpyxl import load_workbook

# Guardrail: allow callers (e.g., ensure_daily_telegram_file.py refresh) to
# explicitly suppress Telegram sends to guarantee exactly-one delivery.
_disable = str(os.environ.get("GA_DISABLE_TELEGRAM_SEND", "")).strip().lower()
if _disable not in ("", "0", "false", "no", "off"):
    print("[TELEGRAM] GA_DISABLE_TELEGRAM_SEND is set; skipping Telegram send.")
    raise SystemExit(0)

# Dedup guard: so envia se houve mudanca real (data nova, genome novo, ou
# fit >= last + 0.05). Override via GA_FORCE_TELEGRAM_SEND=1.
try:
    from analysis.telegram_dedup import should_send_telegram, record_telegram_sent
    _should, _dedup_reason = should_send_telegram()
    if not _should:
        print(f"[TELEGRAM] SKIP (dedup): {_dedup_reason}")
        print("[TELEGRAM] Use GA_FORCE_TELEGRAM_SEND=1 para forcar envio.")
        raise SystemExit(0)
    print(f"[TELEGRAM] dedup OK: {_dedup_reason}")
except SystemExit:
    raise
except Exception as _e:
    # Nao-fatal: se guard falhar, mantem comportamento antigo (sempre envia)
    print(f"[TELEGRAM] WARN: dedup guard falhou, enviando mesmo assim: {_e}")
    record_telegram_sent = None  # noqa: F811 — sinalizador para fim do script

# Use current env token if set, otherwise fallback to the one in ga_run.py
if not os.environ.get("TELEGRAM_BOT_TOKEN"):
    from ga_run import TELEGRAM_BOT_TOKEN as _default_token
    os.environ["TELEGRAM_BOT_TOKEN"] = _default_token

from ga_run import _send_telegram

xlsx = "summary_latest.xlsx"
if not os.path.exists(xlsx):
    raise RuntimeError(f"{xlsx} not found")

# Read Apply + Summary from xlsx
wb = load_workbook(xlsx, read_only=True)

# -- Apply metrics
ws_app = wb["Apply"]
h_app = [c.value for c in next(ws_app.rows)]
rows_app = list(ws_app.iter_rows(min_row=2, values_only=True))

def col(h, name):
    return h.index(name) if name in h else -1

sig_col = col(h_app, "signal")
date_col = col(h_app, "Date")
rec_col = col(h_app, "recomendacao")
pot_col = col(h_app, "potencial_pct")
conf_col = col(h_app, "confidence")
rank_col = col(h_app, "rank")
tkr_col = col(h_app, "ticker")
alvo_col = col(h_app, "alvo")
ent_col = col(h_app, "entrada")

n_buy = sum(1 for r in rows_app if r[sig_col] == "buy")
n_sell = sum(1 for r in rows_app if r[sig_col] == "sell")
n_hold = sum(1 for r in rows_app if r[sig_col] == "hold")
n_total = len(rows_app)
latest_date = max((str(r[date_col]) for r in rows_app), default="N/A")

confs = [float(r[conf_col]) for r in rows_app if r[conf_col] is not None]
conf_med = float(np.median(confs)) if confs else 0.0

# -- Summary metrics
ws_sum = wb["Summary"]
h_sum = [c.value for c in next(ws_sum.rows)]
rows_sum = list(ws_sum.iter_rows(min_row=2, values_only=True))

ret_col_s = col(h_sum, "test_return")
bh_col_s = col(h_sum, "buy_hold_return")
wr_col_s = col(h_sum, "test_win_rate")
mdd_col_s = col(h_sum, "test_mdd")

# Build full arrays with all metrics aligned (so we can filter together)
full_data = []
for r in rows_sum:
    if r[ret_col_s] is None or r[wr_col_s] is None:
        continue
    full_data.append({
        "ret": float(r[ret_col_s]),
        "bh": float(r[bh_col_s]) if r[bh_col_s] is not None else 0.0,
        "wr": float(r[wr_col_s]),
        "mdd": float(r[mdd_col_s]) if r[mdd_col_s] is not None else 0.0,
    })

# Overall WR/MDD (todos os tickers)
all_wrs = np.array([d["wr"] for d in full_data])
all_mdds = np.array([d["mdd"] for d in full_data])
wr_med = float(np.median(all_wrs)) * 100
mdd_med = float(np.median(all_mdds)) * 100

# Filtro WR > 65% para Ret/BH
filtered = [d for d in full_data if d["wr"] > 0.65]
n_filtered = len(filtered)

# Annualized (15 years reference)
n_years = 15.0
def annualize(r):
    return (max(1.0 + r, 0.001)) ** (1.0 / n_years) - 1.0

if n_filtered > 0:
    rets = np.array([d["ret"] for d in filtered])
    bhs = np.array([d["bh"] for d in filtered])
    ret_ann = np.array([annualize(r) for r in rets])
    bh_ann = np.array([annualize(r) for r in bhs])
    alpha_ann = ret_ann - bh_ann

    ret_ann_med = float(np.median(ret_ann)) * 100
    bh_ann_med = float(np.median(bh_ann)) * 100
    alpha_ann_med = float(np.median(alpha_ann)) * 100
    pct_beat_bh = float((rets > bhs).mean()) * 100
else:
    ret_ann_med = 0
    bh_ann_med = 0
    alpha_ann_med = 0
    pct_beat_bh = 0

# -- Win-rate breakdown: alvo hits vs other profitable exits --
# Mostra quantos trades realmente alcancaram o alvo vs foram fechados por
# trailing stop ou time stop com lucro. Usa analysis/win_rate_deep_dive.json.
wr_breakdown_caption = ""
wr_json_path = "analysis/win_rate_deep_dive.json"
if os.path.exists(wr_json_path):
    try:
        with open(wr_json_path, "r", encoding="utf-8") as _fwr:
            _wr_data = json.load(_fwr)
        _wr_med = _wr_data.get("win_rate_median", 0)
        _wr_tgt_med = _wr_data.get("win_rate_target_median", 0)
        _pct_take = _wr_data.get("pct_exit_take_median", 0)
        _pct_stop = _wr_data.get("pct_exit_stop_median", 0)
        wr_breakdown_caption = (
            f"\n\nWR breakdown (full history):\n"
            f"  WR total: {_wr_med:.1f}% | WR alvo: {_wr_tgt_med:.1f}%\n"
            f"  Exits: alvo {_pct_take:.0f}% | stop {_pct_stop:.0f}%"
        )
    except Exception as _e:
        print(f"[WARN] wr_deep_dive load failed: {_e}")

# -- Premium/High-Quality subsets: use CURRENT operational tiers from Summary --
premium_caption = ""
uni_col_s = col(h_sum, "universe")
cagr_col_s = col(h_sum, "test_cagr")
alpha_col_s = col(h_sum, "test_alpha_ann")
if uni_col_s >= 0 and cagr_col_s >= 0 and alpha_col_s >= 0:
    try:
        def _subset_stats(tier_name):
            rows = [
                r for r in rows_sum
                if r[uni_col_s] == tier_name
                and r[wr_col_s] is not None
                and r[cagr_col_s] is not None
                and r[alpha_col_s] is not None
                and r[mdd_col_s] is not None
            ]
            if not rows:
                return None
            return {
                "n": len(rows),
                "wr": float(np.median([float(r[wr_col_s]) for r in rows])) * 100,
                "ret": float(np.median([float(r[cagr_col_s]) for r in rows])) * 100,
                "alp": float(np.median([float(r[alpha_col_s]) for r in rows])) * 100,
                "mdd": float(np.median([float(r[mdd_col_s]) for r in rows])) * 100,
            }

        _prem = _subset_stats("PREMIUM")
        _hq = _subset_stats("HIGH_QUAL")

        subset_lines = []
        if _hq:
            subset_lines.append(
                f"\n\nHigh-Quality atual (n={_hq['n']}):\n"
                f"  WR {_hq['wr']:.1f}% | Ret {_hq['ret']:+.1f}%a.a. "
                f"| Alpha {_hq['alp']:+.1f}pp | MDD {_hq['mdd']:.1f}%"
            )
        if _prem:
            subset_lines.append(
                f"\nPremium atual (n={_prem['n']}):\n"
                f"  WR {_prem['wr']:.1f}% | Ret {_prem['ret']:+.1f}%a.a. "
                f"| Alpha {_prem['alp']:+.1f}pp | MDD {_prem['mdd']:.1f}%"
            )
        premium_caption = "".join(subset_lines)
    except Exception as _e:
        print(f"[WARN] premium/current-tier caption build failed: {_e}")

# Current checkpoint fitness
ck = json.load(open("global_ga_checkpoint.json"))
global_fit = float(ck.get("fitness", 0.0))

# Top BUYS por rank — usa a coluna 'recomendacao' (texto rico com modo de
# execucao, stage, stop, cancelar) em vez de so ticker+preco.
buys = []
for r in rows_app:
    if r[sig_col] != "buy":
        continue
    buys.append({
        "ticker": str(r[tkr_col]),
        "rank": float(r[rank_col]) if r[rank_col] is not None else 0.0,
        "pot": str(r[pot_col]),
        "ent": float(r[ent_col]) if r[ent_col] is not None else 0.0,
        "alvo": float(r[alvo_col]) if r[alvo_col] is not None else 0.0,
        "rec": str(r[rec_col]) if rec_col >= 0 and r[rec_col] is not None else "",
    })
buys.sort(key=lambda x: -x["rank"])

top_buys = ""
if buys:
    # Formato compacto: 1 linha por ticker com TAG + modo de execucao da
    # coluna `recomendacao`, extraindo so o essencial (antes do 1o asterisco).
    # Limite Telegram caption = 1024 chars; cabe ~7 BUYs.
    def _rec_essence(rec: str) -> str:
        """Extrai '[TAG] Fechou@X | MODO | Contexto' compactado do recomendacao.

        Formato atual (post-codex port): linhas separadas por \\n, comecando
        opcionalmente com TAG (PREMIUM/ALTA QUALIDADE), depois 'Fechou @X.XX',
        depois MODO (ex: 'ABAIXO DO PIVOT...'), depois _detalhe, depois
        'Contexto: S2 ...'.

        Estrategia: pega TAG + Fechou + MODO + stage compactado. Descarta
        _detalhe (redundante) e segmentos finais (alvo/stop/cancelar) ja
        visiveis no header do BUY (ticker @ent->alvo).
        """
        if not rec:
            return ""
        rec = rec.strip()
        # split por newline primeiro (formato atual), fallback para asterisks
        if "\n" in rec:
            parts = [p.strip() for p in rec.split("\n") if p.strip()]
        else:
            parts = [p.strip() for p in rec.split("*") if p.strip()]
        if not parts:
            return rec[:60]

        tag = ""
        fechou = ""
        modo = ""
        contexto = ""
        for p in parts:
            p_lower = p.lower()
            if p in ("PREMIUM", "ALTA QUALIDADE") and not tag:
                tag = p
            elif p_lower.startswith("fechou ") and not fechou:
                fechou = p
            elif p_lower.startswith("contexto:") and not contexto:
                contexto = p
            elif not modo and (
                any(k in p_lower for k in (
                    "pivot", "pullback", "esticado", "ordem limite",
                    "compra acionada", "vender", "aguardar", "nao abrir",
                    "breakout", "observar", "abaixo do pivot",
                    "confirmacao", "retomada"))
            ):
                modo = p

        # Compacta contexto: "Contexto: S2 Alta momentum ..." -> "S2 mom"
        stage_short = contexto.replace("Contexto:", "").strip()
        stage_short = (stage_short
                       .replace("S2 Alta ", "S2 ")
                       .replace("momentum", "mom")
                       .replace("pullback", "pull")
                       .replace("atrasado", "late")
                       .replace("S1 Base lateral", "S1 base")
                       .replace("Stage 1", "S1")
                       .replace("Stage 2", "S2")
                       .replace("Stage 3", "S3")
                       .replace("Stage 4", "S4"))
        # primeiro '(' corta parenteses ("(entre Pivot e R1)" etc)
        if "(" in stage_short:
            stage_short = stage_short.split("(")[0].strip()

        # Monta linha final compacta
        segs = []
        if tag: segs.append(tag)
        if fechou: segs.append(fechou)
        if modo:
            # trunca modo se muito longo
            if len(modo) > 48:
                modo = modo[:45] + "..."
            segs.append(modo)
        if stage_short: segs.append(stage_short)
        out = " | ".join(segs) if segs else rec[:60]
        return out[:115]

    lines = []
    for b in buys[:5]:
        # header usa 'entrada' (preco sugerido de compra/ordem limite). O close
        # real do dia esta no essence abaixo ('Fechou @X.XX'), parseado do
        # texto completo da coluna recomendacao.
        header = f"  {b['ticker']} entry@{b['ent']:.2f}->{b['alvo']:.2f} ({b['pot']})"
        essence = _rec_essence(b.get("rec", ""))
        # caption limit Telegram = 1024 chars. Manter essence <= 90 chars.
        if essence and len(essence) > 90:
            essence = essence[:87] + "..."
        if essence:
            lines.append(f"{header}\n    {essence}")
        else:
            lines.append(header)
    top_buys = "\n\nTop BUY:\n" + "\n".join(lines)

wb.close()

# Optional prefix for identification (ex: "Pos merge", "Dev run", "Backfill").
# Set via env var GA_TELEGRAM_PREFIX. Appears as first line da caption.
_prefix = os.environ.get("GA_TELEGRAM_PREFIX", "").strip()
_prefix_line = f"[{_prefix}]\n" if _prefix else ""

caption = (
    f"{_prefix_line}"
    f"Trading Signals B3 - {latest_date}\n"
    f"\n"
    f"Sinais: {n_buy} BUY | {n_hold} HOLD | {n_sell} SELL\n"
    f"Tickers: {n_total}\n"
    f"\n"
    f"WR: {wr_med:.1f}% | MDD: {mdd_med:.1f}%\n"
    f"(WR>65%, n={n_filtered})\n"
    f"Ret a.a.: {ret_ann_med:+.1f}% vs B&H {bh_ann_med:+.1f}% (~{int(n_years)}a)\n"
    f"Alpha a.a.: {alpha_ann_med:+.1f}pp | Batem B&H: {pct_beat_bh:.0f}%\n"
    f"Confidence: {conf_med:.0f} | Fitness: {global_fit:.2f}"
    f"{wr_breakdown_caption}"
    f"{premium_caption}"
    f"{top_buys}"
)

print("=== CAPTION ===")
print(caption)
print()
print(f"=== SENDING {xlsx} ({os.path.getsize(xlsx)} bytes) ===")
ok = _send_telegram(xlsx, caption)
if not ok:
    raise SystemExit(2)

# Registra envio bem-sucedido no tracker para dedup futuro
if record_telegram_sent is not None:
    try:
        record_telegram_sent()
        print("[TELEGRAM] tracker atualizado")
    except Exception as _e:
        print(f"[TELEGRAM] WARN: falha ao atualizar tracker: {_e}")
