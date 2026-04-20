#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Refresh the `close` and `Date` columns of summary_latest.xlsx with today's
live yfinance closes.

Why: stock_etl.py applies SHIFT_FEATURES=1 (see stock_etl.py:772-773) which
shifts the ENTIRE indicator DataFrame — including the OHLCV base columns —
by one bar to avoid feature-leakage during training. That shift is correct
for feature derivation but leaks into the Apply sheet's `close` column:
Date=T carries close from T-1. End users reading the sheet see a stale
close that doesn't match any broker quote.

This module patches only the `close` and `Date` cells of the Apply sheet
using yf.download. Other columns (entrada / alvo / stop / pivot /
potencial_pct / recomendacao) are GA-computed from the same shifted bar and
remain internally consistent — they are forward-looking levels for the next
trading session, which is what the daily signal is for. We do not try to
recompute them here.

Safe to run standalone or as a pipeline step. If yfinance is unreachable or
returns no data for a ticker, that row is left untouched and the issue is
logged.
"""
from __future__ import annotations

import json
import os
import re
import sys
import traceback
from typing import Any

import pandas as pd
from openpyxl import load_workbook

try:
    import yfinance as yf
except ImportError as e:
    raise ImportError("yfinance is required for refresh_close") from e


DEFAULT_XLSX = "summary_latest.xlsx"
STATS_JSON = "analysis/refresh_close_last.json"

# Matches GA's buy execution-mode prefix (see ga_run.py recomendacao block).
# Cobre AMBOS o formato antigo (COMPRAR JA / AGUARDAR ROMPER PIVOT / ORDEM LIMITE)
# e o novo formato codex-port (COMPRA ACIONADA / ABAIXO DO PIVOT / RETOMADA APOS
# PULLBACK / PULLBACK ESTICADO / PULLBACK EM ALTA ordem limite / PULLBACK EM ALTA entrada).
_MODE_RE = re.compile(
    r"(COMPRAR JA|COMPRA ACIONADA|ORDEM LIMITE|"
    r"AGUARDAR ROMPER PIVOT|ABAIXO DO PIVOT:\s*esperar confirmacao|"
    r"RETOMADA APOS PULLBACK:\s*esperar Pivot|"
    r"PULLBACK ESTICADO:\s*preferir recuo|"
    r"PULLBACK EM ALTA:\s*(?:ordem limite|entrada))\s*@\s*\d+(?:[.,]\d+)?"
)

# Matches "Fechou @X.XX" (primeira linha do recomendacao apos port codex).
# Essa linha mostra o close_val do momento de avaliacao da GA — se o close
# foi refrescado, essa string fica stale e precisa atualizar.
_FECHOU_RE = re.compile(r"Fechou\s*@\s*\d+(?:[.,]\d+)?")


def _new_mode(close: float | None, entrada: float | None,
              pivot: float | None) -> str | None:
    """Return the correct buy execution-mode prefix for the current close.

    Conservador: usa modes do formato NOVO (codex port). refresh_close roda
    APOS ga_run, entao o recomendacao sempre tem formato novo.
    """
    if close is None or entrada is None or pivot is None:
        return None
    if close < pivot:
        return f"ABAIXO DO PIVOT: esperar confirmacao @{pivot:.2f}"
    if entrada < close * 0.998:
        return f"ORDEM LIMITE @{entrada:.2f}"
    return f"COMPRA ACIONADA @{entrada:.2f}"


def _rebuild_close_info(recomendacao: str, close: float | None) -> str:
    """Atualiza linha 'Fechou @X.XX' com o novo close. Idempotente se close e None."""
    if not recomendacao or close is None:
        return recomendacao
    new = f"Fechou @{close:.2f}"
    return _FECHOU_RE.sub(new, recomendacao, count=1)


def _rebuild_buy_mode(recomendacao: str, close: float | None,
                      entrada: float | None, pivot: float | None) -> str:
    """Swap the buy execution-mode prefix in-place. Leaves non-buy text alone."""
    if not recomendacao:
        return recomendacao
    new = _new_mode(close, entrada, pivot)
    if new is None:
        return recomendacao
    m = _MODE_RE.search(recomendacao)
    if not m:
        return recomendacao
    return recomendacao[:m.start()] + new + recomendacao[m.end():]


def _coerce(v: Any) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _write_stats(stats: dict) -> None:
    try:
        os.makedirs(os.path.dirname(STATS_JSON), exist_ok=True)
        with open(STATS_JSON, "w", encoding="utf-8") as f:
            json.dump(stats, f, ensure_ascii=False, indent=2, default=str)
    except Exception as e:
        print(f"[refresh_close] WARN: failed to write stats: {e}")


def _fetch_closes(tickers: list[str], period: str = "10d") -> dict[str, tuple[str, float]]:
    """Return {ticker: (iso_date, close)} for each ticker's most recent bar."""
    out: dict[str, tuple[str, float]] = {}
    if not tickers:
        return out
    try:
        data = yf.download(
            tickers,
            period=period,
            group_by="ticker",
            auto_adjust=False,
            progress=False,
            threads=True,
        )
    except Exception as e:
        print(f"[refresh_close] yf.download failed: {e}")
        return out

    if data is None or data.empty:
        return out

    def _extract(ticker: str) -> tuple[str, float] | None:
        try:
            if isinstance(data.columns, pd.MultiIndex):
                if ticker not in data.columns.get_level_values(0):
                    return None
                ss = data[ticker]["Close"].dropna()
            else:
                if len(tickers) != 1:
                    return None
                ss = data["Close"].dropna()
            if len(ss) == 0:
                return None
            last_ts = ss.index[-1]
            last_date = pd.to_datetime(last_ts).date().isoformat()
            return last_date, float(ss.iloc[-1])
        except Exception:
            return None

    for t in tickers:
        rec = _extract(t)
        if rec is not None:
            out[t] = rec
    return out


def refresh(xlsx_path: str = DEFAULT_XLSX) -> dict:
    """Refresh close/Date in xlsx Apply sheet from yfinance. Returns stats dict."""
    stats: dict = {
        "enabled": False,
        "xlsx": xlsx_path,
        "n_rows": 0,
        "n_refreshed": 0,
        "n_missing": 0,
        "max_new_date": None,
        "reason": None,
    }
    if not os.path.exists(xlsx_path):
        stats["reason"] = "xlsx_missing"
        _write_stats(stats)
        return stats

    try:
        wb = load_workbook(xlsx_path)
    except Exception as e:
        stats["reason"] = f"xlsx_load_error:{e}"
        _write_stats(stats)
        return stats

    if "Apply" not in wb.sheetnames:
        stats["reason"] = "no_apply_sheet"
        _write_stats(stats)
        wb.close()
        return stats

    ws = wb["Apply"]
    header = [c.value for c in ws[1]]
    for needed in ("ticker", "close", "Date"):
        if needed not in header:
            stats["reason"] = f"missing_column:{needed}"
            _write_stats(stats)
            wb.close()
            return stats
    i_t = header.index("ticker") + 1
    i_c = header.index("close") + 1
    i_d = header.index("Date") + 1

    def _opt(name: str) -> int:
        return header.index(name) + 1 if name in header else -1

    i_sig = _opt("signal")
    i_sig_pre = _opt("signal_pre_guide")
    i_rec = _opt("recomendacao")
    i_rec_pre = _opt("recomendacao_pre_guide")
    i_ent = _opt("entrada")
    i_piv = _opt("pivot")

    tickers_by_row: list[tuple[int, str]] = []
    unique_tickers: list[str] = []
    seen: set[str] = set()
    for r in range(2, ws.max_row + 1):
        t = ws.cell(row=r, column=i_t).value
        if not t:
            continue
        ticker = str(t).strip().upper()
        tickers_by_row.append((r, ticker))
        if ticker not in seen:
            seen.add(ticker)
            unique_tickers.append(ticker)

    stats["n_rows"] = len(tickers_by_row)
    if not unique_tickers:
        stats["reason"] = "no_tickers"
        _write_stats(stats)
        wb.close()
        return stats

    print(f"[refresh_close] fetching {len(unique_tickers)} tickers from yfinance...")
    closes = _fetch_closes(unique_tickers)
    print(f"[refresh_close] got closes for {len(closes)}/{len(unique_tickers)} tickers")

    refreshed = 0
    missing = 0
    rebuilt_modes = 0
    max_date = None
    for r, ticker in tickers_by_row:
        rec = closes.get(ticker)
        if rec is None:
            missing += 1
            continue
        new_date, new_close = rec
        ws.cell(row=r, column=i_c, value=round(new_close, 4))
        ws.cell(row=r, column=i_d, value=new_date)
        refreshed += 1
        if max_date is None or new_date > max_date:
            max_date = new_date

        # Rebuild recomendacao text:
        # - 'Fechou @X.XX' para TODOS (buy/hold/sell) — reflete novo close
        # - Execution-mode prefix (COMPRA ACIONADA/ORDEM LIMITE/ABAIXO DO PIVOT/etc)
        #   APENAS para buy e se close cruzou regime vs entrada/pivot
        if i_rec < 0:
            continue

        raw_sig = ""
        if i_sig_pre > 0:
            raw_sig = str(ws.cell(row=r, column=i_sig_pre).value or "")
        if not raw_sig.strip() and i_sig > 0:
            raw_sig = str(ws.cell(row=r, column=i_sig).value or "")
        is_buy = (raw_sig.strip().lower() == "buy")

        entrada = pivot = None
        if is_buy and i_ent > 0 and i_piv > 0:
            entrada = _coerce(ws.cell(row=r, column=i_ent).value)
            pivot = _coerce(ws.cell(row=r, column=i_piv).value)

        for col_idx in (i_rec, i_rec_pre):
            if col_idx <= 0:
                continue
            old = ws.cell(row=r, column=col_idx).value
            if not old:
                continue
            # 1. Fechou @X.XX update — all rows (including hold/sell)
            new_rec = _rebuild_close_info(str(old), new_close)
            # 2. Buy execution-mode rebuild (only if signal=buy and pivot/entrada known)
            if is_buy:
                new_rec = _rebuild_buy_mode(new_rec, new_close, entrada, pivot)
            if new_rec != old:
                ws.cell(row=r, column=col_idx, value=new_rec)
                if col_idx == i_rec:
                    rebuilt_modes += 1

    try:
        wb.save(xlsx_path)
    except Exception as e:
        stats["reason"] = f"xlsx_save_error:{e}"
        print(f"[refresh_close] save failed: {e}")
        wb.close()
        _write_stats(stats)
        return stats
    wb.close()

    stats.update({
        "enabled": True,
        "n_refreshed": refreshed,
        "n_missing": missing,
        "n_rebuilt_buy_modes": rebuilt_modes,
        "max_new_date": max_date,
    })
    _write_stats(stats)
    print(f"[refresh_close] OK: refreshed={refreshed} missing={missing} "
          f"modes_rebuilt={rebuilt_modes} max_date={max_date}")
    return stats


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    result = refresh(path)
    print(json.dumps(result, indent=2, default=str))
    sys.exit(0 if result.get("enabled") else 2)
