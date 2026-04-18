#!/usr/bin/env python3
"""
Guarantee that Telegram receives `summary_latest.xlsx` with the latest local trading-day data.

Design goals (automation-safe):
- Prefer *date freshness* over "model freshness": only rebuild the workbook when it's missing/stale
  for the latest store day (or an explicitly requested --date).
- Never double-send: refresh first (if needed), then send exactly once.
- Avoid noisy stacktraces from best-effort refresh helpers.
"""
import argparse
import os
import subprocess
import sys
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
SUMMARY_XLSX = ROOT / "summary_latest.xlsx"
STORE_DIR = ROOT / "output" / "ga_memmap"
CHECKPOINT_JSON = ROOT / "global_ga_checkpoint.json"


def _store_latest_day():
    """
    Best-effort "latest available data day" from the local memmap store.

    This is used as a fallback when refresh is impossible (e.g., no network).
    """
    if not STORE_DIR.exists():
        return None
    try:
        from payload_store import PayloadStore
    except Exception:
        return None

    try:
        store = PayloadStore(str(STORE_DIR), mode="r")
    except Exception:
        return None

    try:
        latest = store.max_date
        return latest.date() if latest is not None else None
    finally:
        try:
            store.close()
        except Exception:
            pass


def _parse_excel_latest_date(xlsx_path: Path):
    if not xlsx_path.exists():
        return None
    wb = load_workbook(xlsx_path, read_only=True)
    try:
        if "Apply" not in wb.sheetnames:
            return None
        ws = wb["Apply"]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return None
        header = list(header)
        if "Date" not in header:
            return None
        date_idx = header.index("Date")
        latest = None
        for row in rows:
            if row is None or date_idx >= len(row):
                continue
            value = row[date_idx]
            if value is None:
                continue
            ts = pd.to_datetime(value, errors="coerce")
            if pd.isna(ts):
                continue
            ts = ts.tz_localize(None) if getattr(ts, "tzinfo", None) is not None else ts
            latest = ts.date() if latest is None or ts.date() > latest else latest
        return latest
    finally:
        wb.close()


def _target_date(raw_date: str | None, store_latest: date | None = None):
    """
    Pick the "trading day" we should guarantee in the workbook.

    - If --date is provided, that is authoritative.
    - Otherwise, prefer the latest day available in the local store, which
      naturally handles weekends/holidays and offline environments.
    - Fall back to local-today only when the store is unavailable.
    """
    if raw_date:
        parsed = pd.to_datetime(raw_date, errors="raise")
        if pd.isna(parsed):
            raise ValueError(f"Invalid --date value: {raw_date}")
        parsed = parsed.tz_localize(None) if getattr(parsed, "tzinfo", None) is not None else parsed
        return parsed.date()
    if store_latest:
        return store_latest
    return pd.Timestamp.now().date()


def _has_current_day_data(xlsx_path: Path, target_day: date):
    latest = _parse_excel_latest_date(xlsx_path)
    return bool(latest == target_day), latest


def _run_predict_daily(full_mode: bool):
    cmd = [sys.executable, "predict_daily.py"]
    if full_mode:
        cmd.append("--full")
    env = os.environ.copy()
    # Avoid double-send: refresh data first, then send exactly once below.
    env["GA_DISABLE_TELEGRAM_SEND"] = "1"
    subprocess.run(cmd, cwd=ROOT, env=env, check=True)


def _run_ga_load_only():
    """
    Rebuild summary_latest.xlsx using the current local store + checkpoint,
    without attempting ETL. This is safe for weekends and offline runs.
    """
    env = os.environ.copy()
    env["GA_RUN_MODE"] = "load"
    env["GA_DISABLE_TELEGRAM_SEND"] = "1"
    # Wrap in a try/except to prevent giant tracebacks from bubbling into the automation logs.
    wrapper = (
        "import sys\n"
        "try:\n"
        "  import ga_run as g\n"
        "  g.run()\n"
        "except Exception as e:\n"
        "  print(f'[WARN] ga_run(load) failed: {e!r}')\n"
        "  sys.exit(1)\n"
    )
    subprocess.run([sys.executable, "-c", wrapper], cwd=ROOT, env=env, check=True)


def _run_send_telegram():
    subprocess.run([sys.executable, "send_telegram.py"], cwd=ROOT, check=True)


def main():
    parser = argparse.ArgumentParser(description="Ensure Telegram receives summary_latest.xlsx with current-day data.")
    parser.add_argument("--date", type=str, default=None, help="Target date in YYYY-MM-DD format. Defaults to local today.")
    parser.add_argument("--no-refresh", action="store_true", help="Do not run predict_daily.py if summary_latest.xlsx is stale.")
    parser.add_argument("--full", action="store_true", help="Use predict_daily.py --full when refresh is needed.")
    parser.add_argument(
        "--allow-etl-refresh",
        action="store_true",
        help="Allow running predict_daily.py (ETL/network) when workbook is stale and load-only refresh fails.",
    )
    parser.add_argument(
        "--refresh-model",
        action="store_true",
        help="If checkpoint is newer than the workbook, attempt a load-only rebuild even if the workbook date is fresh.",
    )
    parser.add_argument("--dry-run", action="store_true", help="Check freshness without sending Telegram.")
    args = parser.parse_args()

    store_latest = _store_latest_day()
    target_day = _target_date(args.date, store_latest)
    if store_latest:
        print(f"[CHECK] store_latest_day={store_latest.isoformat()}", flush=True)

    fresh, latest = _has_current_day_data(SUMMARY_XLSX, target_day)
    print(f"[CHECK] target_day={target_day.isoformat()} latest_in_xlsx={latest}", flush=True)

    # Optional: model freshness refresh (but only on explicit request).
    if (
        args.refresh_model
        and not args.no_refresh
        and CHECKPOINT_JSON.exists()
        and SUMMARY_XLSX.exists()
        and CHECKPOINT_JSON.stat().st_mtime > SUMMARY_XLSX.stat().st_mtime + 1.0
    ):
        print("[REFRESH] Checkpoint newer than workbook; attempting GA load-only rebuild.", flush=True)
        try:
            _run_ga_load_only()
            fresh, latest = _has_current_day_data(SUMMARY_XLSX, target_day)
            print(f"[CHECK] after refresh latest_in_xlsx={latest}", flush=True)
        except subprocess.CalledProcessError as exc:
            print(f"[WARN] GA load-only rebuild failed: {exc}", flush=True)

    if not fresh:
        if args.no_refresh:
            raise SystemExit(f"summary_latest.xlsx is stale or missing (latest={latest}, target={target_day})")

        print("[REFRESH] summary_latest.xlsx is stale or missing; attempting GA load-only rebuild.", flush=True)
        refresh_ok = False
        try:
            _run_ga_load_only()
            refresh_ok = True
        except subprocess.CalledProcessError as exc:
            print(f"[WARN] GA load-only rebuild failed: {exc}", flush=True)

        if (not refresh_ok) and args.allow_etl_refresh:
            print("[REFRESH] Falling back to predict_daily.py (ETL/network) because --allow-etl-refresh is set.", flush=True)
            try:
                _run_predict_daily(full_mode=args.full)
                refresh_ok = True
            except subprocess.CalledProcessError as exc:
                print(f"[WARN] predict_daily.py failed: {exc}", flush=True)

        if refresh_ok:
            fresh, latest = _has_current_day_data(SUMMARY_XLSX, target_day)
            print(f"[CHECK] after refresh latest_in_xlsx={latest}", flush=True)
        elif store_latest:
            # Network-constrained fallback: send the most recent day the local store supports.
            fallback_fresh, fallback_latest = _has_current_day_data(SUMMARY_XLSX, store_latest)
            if fallback_fresh:
                print(
                    "[FALLBACK] Unable to refresh current-day data; "
                    f"using store_latest_day={store_latest.isoformat()} for Telegram send."
                )
                target_day = store_latest
                fresh, latest = True, fallback_latest
            else:
                raise SystemExit(
                    "Could not refresh daily data and summary_latest.xlsx does not contain "
                    f"store_latest_day={store_latest.isoformat()} (latest_in_xlsx={fallback_latest})."
                )

    if not fresh:
        raise SystemExit(f"Could not obtain current-day data in summary_latest.xlsx (latest={latest}, target={target_day})")

    if args.dry_run:
        print("[DRY-RUN] Current-day data is available; Telegram send skipped.")
        return

    print("[SEND] Sending summary_latest.xlsx to Telegram.", flush=True)
    try:
        _run_send_telegram()
    except subprocess.CalledProcessError as exc:
        print(f"[ERROR] Telegram send failed (exit={exc.returncode}).")
        raise SystemExit(2)


if __name__ == "__main__":
    main()
