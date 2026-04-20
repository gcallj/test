#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Apply Brazilian stock-guide PDF data as a post-GA overlay on summary_latest.xlsx.

Soft-adjust rules (at most one adjustment per row):
  * GA buy  + guide sell (and confidence < 85)    -> demote to hold, conf -10
  * GA buy  + guide buy                           -> conf +10 (capped at 100)
  * GA hold + guide buy  (65 <= conf < 75) + SAFE -> promote to buy, conf +5
  * GA sell + guide buy                           -> conf -5 (signal kept)
  * GA sell + guide sell                          -> confirm (no change)

Safety guards (block promotion even when conf + guide agree):
  * Stage 3 Topo / Stage 4 Queda -> never promote
  * recomendacao starts with "NAO ABRIR" or "VENDER" -> never promote
  * alvo <= close (no upside) -> never promote

Idempotency: overlay preserves original `signal`, `confidence`, `recomendacao`
in mirror columns (`signal_pre_guide`, `confidence_pre_guide`,
`recomendacao_pre_guide`) on first run, and always re-applies rules from those
pristine values on subsequent runs. Each run produces the same result for a
given xlsx + PDF pair.

When we mutate `signal`, we also rebuild `recomendacao` so the text matches the
new stance, mirroring the ga_run template (COMPRAR JA / ORDEM LIMITE /
AGUARDAR ROMPER PIVOT for buys; NAO ABRIR / AGUARDAR for holds). For rows we
don't mutate, we prepend a `[GUIA ...]` bracket tag so the user can see whether
the guide agreed, disagreed, or was silent — following the existing
`[PREMIUM]` / `[BAIXA QUAL]` annotation convention in ga_run.py.

Never fatal: if the PDF is missing, unparseable, or the xlsx can't be opened,
logs a warning and returns {"enabled": False, "reason": "..."}.
"""
from __future__ import annotations

import glob
import json
import os
import re
import traceback
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from guide_pdf_parser import SCHEMA_COLUMNS, parse_guide

DEFAULT_XLSX = "summary_latest.xlsx"
DEFAULT_GUIDES_DIR = "./guides"
STATS_JSON = "analysis/guide_overlay_last.json"

ENRICHMENT_COLUMNS = [
    "guide_rec",
    "guide_target",
    "guide_upside_pct",
    "guide_qual_score",
    "guide_momentum_score",
    "guide_rec_score",
    "guide_report_date",
    "decision_note",
]
MIRROR_COLUMNS = [
    "signal_pre_guide",
    "confidence_pre_guide",
    "recomendacao_pre_guide",
]
NEW_APPLY_COLUMNS = ENRICHMENT_COLUMNS + MIRROR_COLUMNS

# Conservative promotion gate: if recomendacao starts with either of these
# words, the technical picture explicitly rejects opening a long position.
_PROMOTE_VETO_RE = re.compile(r"^\s*(?:\[[^\]]+\]\s*)*(NAO ABRIR|VENDER)\b", re.IGNORECASE)
# Stage phrases that block promotion even when WR/conf would allow it.
_STAGE_VETO_RE = re.compile(r"Stage\s*[34]\b", re.IGNORECASE)

# Upper confidence bound at which guide SELL is strong enough to override a GA BUY.
_DEMOTE_CONF_CEILING = 85.0
# Bounds for "borderline hold" — the only zone where a guide buy can promote.
_PROMOTE_CONF_LO, _PROMOTE_CONF_HI = 65.0, 75.0


def _pick_pdf(env_override: str | None, guides_dir: str) -> str | None:
    if env_override and os.path.exists(env_override):
        return env_override
    env_path = os.environ.get("PDF_GUIDE_PATH")
    if env_path and os.path.exists(env_path):
        return env_path
    if not os.path.isdir(guides_dir):
        return None
    pdfs = sorted(
        glob.glob(os.path.join(guides_dir, "*.pdf")),
        key=lambda p: os.path.getmtime(p),
        reverse=True,
    )
    return pdfs[0] if pdfs else None


def _write_stats(stats: dict) -> None:
    try:
        os.makedirs(os.path.dirname(STATS_JSON), exist_ok=True)
        with open(STATS_JSON, "w", encoding="utf-8") as f:
            json.dump(stats, f, ensure_ascii=False, indent=2, default=str)
    except Exception as e:
        print(f"[guide_overlay] WARN: failed to write stats {STATS_JSON}: {e}")


def _coerce_float(v: Any) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _clamp(v: float, lo: float = 0.0, hi: float = 100.0) -> float:
    return max(lo, min(hi, v))


def _strip_universe_prefix(rec: str) -> tuple[str, str]:
    """Split '[PREMIUM] ORDEM LIMITE...' -> ('[PREMIUM] ', 'ORDEM LIMITE...')."""
    if not rec:
        return "", ""
    m = re.match(r"^((?:\[[^\]]+\]\s*)+)(.*)", rec, re.DOTALL)
    if not m:
        return "", rec
    return m.group(1), m.group(2)


def _stage_short(estagio: str | None) -> str:
    """'Stage 2 Alta - momentum' -> 'S2 Alta - momentum'."""
    if not estagio:
        return ""
    return str(estagio).replace("Stage ", "S")


def _fmt_price(v: float | None) -> str:
    return f"@{v:.2f}" if v is not None and not pd.isna(v) else "@?"


def _build_buy_text(ctx: dict) -> str:
    """Rebuild a buy recomendacao mirroring ga_run.py:4321-4342."""
    close = ctx["close"]
    entrada = ctx["entrada"]
    pivot = ctx["pivot"]
    alvo = ctx["alvo"]
    stop = ctx["stop"]
    cancel = ctx["cancelar"]
    stage = ctx["stage_short"]
    wr = ctx["wr"]
    wr_alvo = ctx["wr_alvo"]
    potencial_pct = ctx["potencial_pct"]
    stop_pct = ctx["stop_pct"]

    if close is not None and pivot is not None and close < pivot:
        modo = f"AGUARDAR ROMPER PIVOT {_fmt_price(pivot)}"
    elif entrada is not None and close is not None and entrada < close * 0.998:
        modo = f"ORDEM LIMITE {_fmt_price(entrada)}"
    else:
        modo = f"COMPRAR JA {_fmt_price(entrada if entrada is not None else close)}"

    pot_str = f"+{potencial_pct:.1f}%" if potencial_pct is not None else "+?"
    wr_str = f"{wr:.0f}%" if wr is not None else "?"
    wralvo_str = f"{wr_alvo:.0f}%" if wr_alvo is not None else "?"
    stop_str = f"-{stop_pct:.1f}%" if stop_pct is not None else "-?%"

    return (
        f"{modo} * {stage} * "
        f"{pot_str} alvo {_fmt_price(alvo)} "
        f"(WR {wr_str} / alvo {wralvo_str}) * "
        f"Stop {_fmt_price(stop)} ({stop_str}) * "
        f"Cancelar se perder {_fmt_price(cancel if cancel is not None else pivot)}"
    )


def _build_hold_text(ctx: dict) -> str:
    """Rebuild a hold recomendacao matching the original stage language."""
    estagio = ctx.get("estagio") or ""
    stage = ctx["stage_short"]
    pivot = ctx["pivot"]
    r1 = ctx.get("r1")
    s1 = ctx.get("s1")

    if "Stage 4" in estagio:
        return f"NAO ABRIR * {stage} * Abaixo MA200 (tendencia de baixa)"
    if "Stage 3" in estagio:
        return f"NAO ABRIR * {stage} * Vender posicoes se perder Pivot {_fmt_price(pivot)}"
    if "Stage 2" in estagio:
        tgt = r1 if r1 is not None else pivot
        return (
            f"AGUARDAR * {stage} * "
            f"Comprar apos romper {_fmt_price(tgt)} ou recuar ate {_fmt_price(s1)} * "
            f"Cancelar se perder {_fmt_price(ctx.get('cancelar') or pivot)}"
        )
    if "Stage 1" in estagio:
        return (
            f"AGUARDAR BREAKOUT {_fmt_price(r1)} * {stage} * "
            f"Evitar se perder {_fmt_price(s1)}"
        )
    return f"OBSERVAR * {stage}"


def _safe_to_promote(ctx: dict, pre_rec: str) -> tuple[bool, str | None]:
    """Decide whether a hold->buy promotion is technically safe.

    Returns (True, None) to promote, or (False, reason) to skip.
    """
    close = ctx.get("close")
    alvo = ctx.get("alvo")
    potencial_pct = ctx.get("potencial_pct")
    estagio = ctx.get("estagio") or ""

    if _PROMOTE_VETO_RE.search(pre_rec or ""):
        return False, "veto_text"
    if _STAGE_VETO_RE.search(estagio):
        return False, "veto_stage"
    if alvo is not None and close is not None and alvo <= close:
        return False, "veto_no_upside"
    if potencial_pct is not None and potencial_pct <= 0:
        return False, "veto_neg_potencial"
    return True, None


def _fresh_upside(target: Any, close: float | None) -> float | None:
    """Upside of the guide target vs TODAY'S close (not the PDF's stale ref)."""
    if target is None or pd.isna(target):
        return None
    if close is None or close <= 0:
        return None
    try:
        return (float(target) / float(close) - 1.0) * 100.0
    except (ValueError, TypeError, ZeroDivisionError):
        return None


def _target_tag_body(label: str, guide: dict, close: float | None) -> str:
    tgt = guide.get("guide_target")
    parts = [label]
    if tgt is not None and not pd.isna(tgt):
        parts.append(f"alvo R$ {float(tgt):.2f}")
    up = _fresh_upside(tgt, close)
    if up is not None:
        parts.append(f"({up:+.1f}%)")
    return " ".join(parts)


def _guide_tag_agree_buy(guide: dict, close: float | None) -> str:
    return f"[{_target_tag_body('GUIA OK COMPRAR', guide, close)}] "


def _guide_tag_agree_sell() -> str:
    return "[GUIA OK VENDER] "


def _guide_tag_conflict_buy_kept_high_conf() -> str:
    return "[GUIA VENDER - CONVICCAO ALTA MANTIDA] "


def _guide_tag_conflict_sell_demoted() -> str:
    return "[GUIA VENDER - BUY DEMOVIDO] "


def _guide_tag_conflict_sell_vs_buy(guide: dict, close: float | None) -> str:
    return f"[{_target_tag_body('GUIA COMPRAR', guide, close)} - DIVERGENCIA] "


def _guide_tag_promoted(guide: dict, close: float | None) -> str:
    return f"[{_target_tag_body('GUIA PROMOVE HOLD-> BUY', guide, close)}] "


def _guide_tag_promote_vetoed(reason: str, guide: dict, close: float | None) -> str:
    return f"[{_target_tag_body('GUIA COMPRAR', guide, close)} - BLOQUEADO {reason}] "


def apply_overlay(
    xlsx_path: str = DEFAULT_XLSX,
    guides_dir: str = DEFAULT_GUIDES_DIR,
    env_override: str | None = None,
) -> dict:
    """Apply guide overlay to the Apply sheet of xlsx_path. See module docstring."""
    stats: dict = {
        "enabled": False,
        "pdf_file": None,
        "tickers_matched": 0,
        "n_agree_buy": 0,
        "n_agree_sell": 0,
        "n_conflict_buy_vs_sell": 0,
        "n_conflict_sell_vs_buy": 0,
        "n_promoted_hold_to_buy": 0,
        "n_promote_vetoed": 0,
        "n_demoted_buy_to_hold": 0,
        "reason": None,
    }

    pdf_path = _pick_pdf(env_override, guides_dir)
    if not pdf_path:
        stats["reason"] = "no_pdf"
        _write_stats(stats)
        return stats
    stats["pdf_file"] = os.path.basename(pdf_path)

    if not os.path.exists(xlsx_path):
        stats["reason"] = f"xlsx_missing:{xlsx_path}"
        _write_stats(stats)
        return stats

    try:
        guide_df = parse_guide(pdf_path)
    except Exception as e:
        stats["reason"] = f"parse_error:{e}"
        print(f"[guide_overlay] parse failed: {e}\n{traceback.format_exc()}")
        _write_stats(stats)
        return stats
    if guide_df.empty:
        stats["reason"] = "parse_empty"
        _write_stats(stats)
        return stats

    guide_by_ticker = {
        str(row["ticker"]): {c: row[c] for c in SCHEMA_COLUMNS if c != "ticker"}
        for _, row in guide_df.iterrows()
    }

    try:
        wb = load_workbook(xlsx_path)
    except Exception as e:
        stats["reason"] = f"xlsx_load_error:{e}"
        print(f"[guide_overlay] xlsx load failed: {e}")
        _write_stats(stats)
        return stats

    if "Apply" not in wb.sheetnames:
        stats["reason"] = "no_apply_sheet"
        _write_stats(stats)
        wb.close()
        return stats

    ws = wb["Apply"]
    header = [c.value for c in ws[1]]

    def col_idx(name: str) -> int:
        return header.index(name) if name in header else -1

    required = ["ticker", "signal", "recomendacao", "confidence", "close",
                "entrada", "pivot", "alvo", "stop", "cancelar", "estagio",
                "win_rate", "wr_alvo", "potencial_pct", "r1", "s1",
                "universe", "signal_ga"]
    for c in ["ticker", "signal"]:
        if col_idx(c) < 0:
            stats["reason"] = f"missing_column:{c}"
            _write_stats(stats)
            wb.close()
            return stats

    idx = {c: col_idx(c) for c in required}

    # Append new columns (idempotent).
    new_col_positions: dict[str, int] = {}
    for name in NEW_APPLY_COLUMNS:
        i = col_idx(name)
        if i < 0:
            i = len(header)
            header.append(name)
            ws.cell(row=1, column=i + 1, value=name)
        new_col_positions[name] = i

    def getv(r: int, name: str):
        ci = idx.get(name, -1)
        return ws.cell(row=r, column=ci + 1).value if ci >= 0 else None

    def setv(r: int, name: str, value):
        ci = idx.get(name, -1)
        if ci >= 0:
            ws.cell(row=r, column=ci + 1, value=value)

    def getn(r: int, name: str):
        return new_col_positions.get(name, -1)

    def getval_new(r: int, name: str):
        ci = new_col_positions.get(name, -1)
        return ws.cell(row=r, column=ci + 1).value if ci >= 0 else None

    def setval_new(r: int, name: str, value):
        ci = new_col_positions.get(name, -1)
        if ci >= 0:
            ws.cell(row=r, column=ci + 1, value=value)

    max_row = ws.max_row
    matched = 0
    agree_buy = agree_sell = conflict_bs = conflict_sb = 0
    promoted = vetoed = demoted = 0

    for r in range(2, max_row + 1):
        ticker = ws.cell(row=r, column=idx["ticker"] + 1).value
        if not ticker:
            continue
        ticker = str(ticker).strip().upper()

        # --- Restore pre-guide state (makes overlay idempotent) ---
        pre_sig = getval_new(r, "signal_pre_guide")
        if pre_sig is None or str(pre_sig).strip() == "":
            pre_sig = getv(r, "signal")
            setval_new(r, "signal_pre_guide", pre_sig)
        pre_conf = getval_new(r, "confidence_pre_guide")
        if pre_conf is None:
            pre_conf = getv(r, "confidence")
            setval_new(r, "confidence_pre_guide", pre_conf)
        pre_rec = getval_new(r, "recomendacao_pre_guide")
        if pre_rec is None or str(pre_rec).strip() == "":
            pre_rec = getv(r, "recomendacao")
            setval_new(r, "recomendacao_pre_guide", pre_rec)

        guide = guide_by_ticker.get(ticker)
        if not guide:
            # Reset to pristine state + clear enrichment fields.
            setv(r, "signal", pre_sig)
            setv(r, "confidence", pre_conf)
            setv(r, "recomendacao", pre_rec)
            for name in ENRICHMENT_COLUMNS:
                setval_new(r, name, None)
            continue

        matched += 1
        g_rec = guide.get("guide_rec")
        g_rec_str = str(g_rec).lower() if g_rec is not None and not (
            isinstance(g_rec, float) and pd.isna(g_rec)) else None

        # Enrichment columns (always populated when a match exists).
        def _gv(k):
            v = guide.get(k)
            if isinstance(v, float) and pd.isna(v):
                return None
            return v

        # guide_upside_pct is overwritten below with a fresh value computed
        # against today's close (the PDF's raw pct is anchored to the PDF's
        # price-at-report-time and would misalign with the live close column).
        setval_new(r, "guide_rec", g_rec_str)
        setval_new(r, "guide_target", _gv("guide_target"))
        setval_new(r, "guide_qual_score", _gv("guide_qual_score"))
        setval_new(r, "guide_momentum_score", _gv("guide_momentum_score"))
        setval_new(r, "guide_rec_score", _gv("guide_rec_score"))
        rep_dt = _gv("guide_report_date")
        setval_new(r, "guide_report_date",
                   rep_dt.isoformat() if hasattr(rep_dt, "isoformat") else rep_dt)

        # Build a row context for text generation + safety checks.
        ctx = {
            "close": _coerce_float(getv(r, "close")),
            "entrada": _coerce_float(getv(r, "entrada")),
            "pivot": _coerce_float(getv(r, "pivot")),
            "alvo": _coerce_float(getv(r, "alvo")),
            "stop": _coerce_float(getv(r, "stop")),
            "cancelar": _coerce_float(getv(r, "cancelar")),
            "r1": _coerce_float(getv(r, "r1")),
            "s1": _coerce_float(getv(r, "s1")),
            "wr": _coerce_float(getv(r, "win_rate")),
            "wr_alvo": _coerce_float(getv(r, "wr_alvo")),
            "estagio": getv(r, "estagio"),
            "stage_short": _stage_short(getv(r, "estagio")),
        }
        pot_raw = getv(r, "potencial_pct")
        if isinstance(pot_raw, str):
            m = re.search(r"-?\d+(?:[.,]\d+)?", pot_raw)
            ctx["potencial_pct"] = float(m.group(0).replace(",", ".")) if m else None
        else:
            ctx["potencial_pct"] = _coerce_float(pot_raw)
        if ctx["stop"] is not None and ctx["entrada"] and ctx["entrada"] > 0:
            ctx["stop_pct"] = abs((ctx["entrada"] - ctx["stop"]) / ctx["entrada"] * 100.0)
        else:
            ctx["stop_pct"] = None

        # Fresh guide upside vs today's close (pre-computed so every consumer
        # — bracket tag, xlsx column, stats — reads the same number).
        fresh_up = _fresh_upside(_gv("guide_target"), ctx["close"])
        setval_new(r, "guide_upside_pct", round(fresh_up, 2) if fresh_up is not None else None)

        # Baseline for this row: restore from pre-guide snapshot.
        pre_sig_str = str(pre_sig or "").strip().lower()
        pre_conf_f = _coerce_float(pre_conf)
        pre_universe_prefix, pre_rec_body = _strip_universe_prefix(str(pre_rec or ""))

        note = None
        new_sig = pre_sig_str
        new_conf = pre_conf_f
        new_rec = pre_rec

        if g_rec_str is None:
            pass
        elif pre_sig_str == "buy" and g_rec_str == "sell":
            conflict_bs += 1
            if pre_conf_f is None or pre_conf_f < _DEMOTE_CONF_CEILING:
                new_sig = "hold"
                note = "guide_conflict_sell"
                if pre_conf_f is not None:
                    new_conf = _clamp(pre_conf_f - 10.0)
                demoted += 1
                new_rec = (pre_universe_prefix + _guide_tag_conflict_sell_demoted()
                           + _build_hold_text(ctx))
            else:
                note = "guide_conflict_sell_kept"
                new_rec = (pre_universe_prefix + _guide_tag_conflict_buy_kept_high_conf()
                           + pre_rec_body)
        elif pre_sig_str == "buy" and g_rec_str == "buy":
            note = "guide_confirm_buy"
            if pre_conf_f is not None:
                new_conf = _clamp(pre_conf_f + 10.0)
            agree_buy += 1
            new_rec = pre_universe_prefix + _guide_tag_agree_buy(guide, ctx["close"]) + pre_rec_body
        elif pre_sig_str == "hold" and g_rec_str == "buy":
            if pre_conf_f is not None and _PROMOTE_CONF_LO <= pre_conf_f < _PROMOTE_CONF_HI:
                safe, reason = _safe_to_promote(ctx, str(pre_rec or ""))
                if safe:
                    new_sig = "buy"
                    note = "guide_promote_buy"
                    new_conf = _clamp(pre_conf_f + 5.0)
                    promoted += 1
                    universe = getv(r, "universe") or "REGULAR"
                    univ_prefix = (f"[{universe}] " if universe != "REGULAR"
                                   else "[REGULAR-cuidado] ")
                    new_rec = (univ_prefix + _guide_tag_promoted(guide, ctx["close"])
                               + _build_buy_text(ctx))
                else:
                    note = f"guide_promote_vetoed:{reason}"
                    vetoed += 1
                    new_rec = (pre_universe_prefix
                               + _guide_tag_promote_vetoed(reason, guide, ctx["close"])
                               + pre_rec_body)
        elif pre_sig_str == "sell" and g_rec_str == "sell":
            note = "guide_confirm_sell"
            agree_sell += 1
            new_rec = pre_universe_prefix + _guide_tag_agree_sell() + pre_rec_body
        elif pre_sig_str == "sell" and g_rec_str == "buy":
            note = "guide_conflict_buy"
            if pre_conf_f is not None:
                new_conf = _clamp(pre_conf_f - 5.0)
            conflict_sb += 1
            new_rec = (pre_universe_prefix + _guide_tag_conflict_sell_vs_buy(guide, ctx["close"])
                       + pre_rec_body)

        # Write adjusted values (always write, so re-runs re-materialize from pre_*).
        setv(r, "signal", new_sig)
        if new_conf is not None:
            setv(r, "confidence", round(new_conf, 1))
        setv(r, "recomendacao", new_rec)
        setval_new(r, "decision_note", note)

    try:
        wb.save(xlsx_path)
    except Exception as e:
        stats["reason"] = f"xlsx_save_error:{e}"
        print(f"[guide_overlay] save failed: {e}")
        wb.close()
        _write_stats(stats)
        return stats
    wb.close()

    stats.update({
        "enabled": True,
        "tickers_matched": matched,
        "n_agree_buy": agree_buy,
        "n_agree_sell": agree_sell,
        "n_conflict_buy_vs_sell": conflict_bs,
        "n_conflict_sell_vs_buy": conflict_sb,
        "n_promoted_hold_to_buy": promoted,
        "n_promote_vetoed": vetoed,
        "n_demoted_buy_to_hold": demoted,
        "reason": None,
    })
    _write_stats(stats)
    return stats


if __name__ == "__main__":
    print(json.dumps(apply_overlay(), indent=2, default=str))
